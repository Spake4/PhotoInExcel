import openpyxl
from openpyxl.styles import PatternFill

rows = 336
cols = 279

matrix = [['' for _ in range(cols)] for _ in range(rows)]

with open(r'd:\PhotoInExcel\src\dataFile.txt', 'r') as file:
    for r in range(rows):
        for c in range(cols):
            char = file.read(1)
            while char == '#' or char == ' ':
                char = file.read(1)
            if char.isdigit():
                num = char
                while True:
                    char = file.read(1)
                    if char.isdigit():
                        num += char
                    else:
                        break
                matrix[r][c] = num

workbook = openpyxl.Workbook()
sheet = workbook.active

purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid') 

for i in range(rows):
    for j in range(cols):
        if matrix[i][j]:
            cell = sheet.cell(row=i+1, column=j+1, value=matrix[i][j])
            cell.fill = purple_fill 
            
workbook.save(r'd:\PhotoInExcel\src\etterna.xlsx')

print("WorkComplite")